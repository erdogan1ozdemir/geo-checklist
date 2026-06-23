#!/usr/bin/env python3
"""
Extract the GEO/AEO checklist workbook into structured JSON for the dashboard.

Classifies every row of the E-Ticaret and Hizmet sheets into:
  title | meta | header | phase | section | task | subtask
and builds a nested phase -> section -> task -> subtask tree, keeping each
row's ABSOLUTE 1-based sheet row index so the export engine can map a UI
selection back to the original template rows.
"""
import json
import os
import openpyxl

SRC = os.path.join(os.path.dirname(__file__), "..", "..", "GEO-AEO-Checklist-v5.xlsx")
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data")

# Column map (1-based) -> field
COLS = {
    2: "faz",
    3: "gorevTipi",
    4: "kanal",
    5: "aksiyon",
    6: "detay",
    7: "arac",
    8: "oncelik",
    9: "sorumlu",
    10: "durum",
    11: "markaNotlari",
    12: "inboundNotlari",
}

SHEET_FILE = {  # sheet name -> worksheets xml file (sheet order in workbook)
    "E-Ticaret": "sheet1.xml",
    "Hizmet": "sheet2.xml",
}


def cellval(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    return str(v).strip()


def row_fields(ws, r):
    return {field: cellval(ws, r, c) for c, field in COLS.items()}


def classify(ws, r):
    a = cellval(ws, r, 1)
    e = cellval(ws, r, 5)
    if r == 1:
        return "title"
    if r == 2:
        return "meta"
    if r == 3:
        return "header"
    # banners: col A holds the title, col B empty
    b = cellval(ws, r, 2)
    has_check = "☐" in a or "☑" in a
    if not has_check and a and not b:
        if a.upper().startswith("FAZ"):
            return "phase"
        return "section"
    # otherwise a task/subtask row
    if e.startswith("→"):
        return "subtask"
    return "task"


def build_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    title = cellval(ws, 1, 1)
    meta = cellval(ws, 2, 1)
    phases = []
    cur_phase = None
    cur_section = None
    cur_task = None

    counts = {"phase": 0, "section": 0, "task": 0, "subtask": 0}

    for r in range(4, ws.max_row + 1):
        kind = classify(ws, r)
        a = cellval(ws, r, 1)
        if kind == "phase":
            cur_phase = {"id": f"{sheet_name}-r{r}", "row": r, "title": a, "sections": []}
            phases.append(cur_phase)
            cur_section = None
            cur_task = None
            counts["phase"] += 1
        elif kind == "section":
            # split leading emoji + "N." from the title text
            cur_section = {"id": f"{sheet_name}-r{r}", "row": r, "title": a, "tasks": []}
            if cur_phase is None:
                cur_phase = {"id": f"{sheet_name}-r0", "row": r, "title": "", "sections": []}
                phases.append(cur_phase)
            cur_phase["sections"].append(cur_section)
            cur_task = None
            counts["section"] += 1
        elif kind == "task":
            f = row_fields(ws, r)
            cur_task = {
                "id": f"{sheet_name}-r{r}",
                "row": r,
                "kind": "task",
                **f,
                "subtasks": [],
            }
            if cur_section is None:
                cur_section = {"id": f"{sheet_name}-r{r}s", "row": r, "title": "", "tasks": []}
                if cur_phase is None:
                    cur_phase = {"id": f"{sheet_name}-r{r}p", "row": r, "title": "", "sections": []}
                    phases.append(cur_phase)
                cur_phase["sections"].append(cur_section)
            cur_section["tasks"].append(cur_task)
            counts["task"] += 1
        elif kind == "subtask":
            f = row_fields(ws, r)
            st = {"id": f"{sheet_name}-r{r}", "row": r, "kind": "subtask", **f}
            if cur_task is None:
                # orphan subtask -> promote to a task so it is never lost
                cur_task = {
                    "id": f"{sheet_name}-r{r}",
                    "row": r,
                    "kind": "task",
                    **f,
                    "subtasks": [],
                }
                if cur_section is None:
                    cur_section = {"id": f"{sheet_name}-orphan", "row": r, "title": "", "tasks": []}
                    if cur_phase is None:
                        cur_phase = {"id": f"{sheet_name}-orphanp", "row": r, "title": "", "sections": []}
                        phases.append(cur_phase)
                    cur_phase["sections"].append(cur_section)
                cur_section["tasks"].append(cur_task)
                counts["task"] += 1
            else:
                cur_task["subtasks"].append(st)
                counts["subtask"] += 1

    return {
        "sheet": sheet_name,
        "sheetFile": SHEET_FILE[sheet_name],
        "title": title,
        "meta": meta,
        "phases": phases,
        "counts": counts,
        "maxRow": ws.max_row,
    }


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)
    combined = {}
    for name in ["E-Ticaret", "Hizmet"]:
        data = build_sheet(wb, name)
        combined[name] = data
        print(f"{name}: {data['counts']} (maxRow={data['maxRow']})")
    out = os.path.join(OUT_DIR, "checklist.json")
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(combined, fh, ensure_ascii=False, indent=1)
    size = os.path.getsize(out)
    print(f"wrote {out} ({size//1024} KB)")


if __name__ == "__main__":
    main()

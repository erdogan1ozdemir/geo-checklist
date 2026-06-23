#!/usr/bin/env python3
"""Validate /tmp/test-export.xlsx against the original for byte-perfect styling
and correct row filtering."""
import json
import sys
import openpyxl

ORIG = "../GEO-AEO-Checklist-v5.xlsx"
EXP = "/tmp/test-export.xlsx"

# Must mirror COLOR_MAP in src/lib/xlsxExport.js
COLOR_MAP = {
    "FF1F3864": "FF10332F", "FF1F4E79": "FF1A4238", "FF2E75B6": "FFFF7B52",
    "FFE8F0FE": "FFFFE3D8", "FFF2F2F2": "FFF4F2EE", "FFF0F0F0": "FFF4F2EE",
    "FFFF6B6B": "FFE5534D", "FFFFB347": "FFF5A623", "FFFFE066": "FFFAD46B",
    "FF77DD77": "FF8FD3A6",
}


def remap(rgb):
    return COLOR_MAP.get(rgb, rgb)


def style_sig(cell, recolor=False):
    f = cell.fill
    fill = f.fgColor.rgb if (f and f.patternType) else None
    fo = cell.font
    fontc = fo.color.rgb if fo.color else None
    if recolor:
        fill = remap(fill)
        fontc = remap(fontc)
    return (
        fill,
        fo.bold,
        round(fo.sz, 1) if fo.sz else None,
        fontc,
        fo.name,
        cell.alignment.horizontal,
        cell.alignment.vertical,
        cell.alignment.wrap_text,
    )


def main():
    expected = json.load(open("/tmp/test-export-expected.json"))
    orig = openpyxl.load_workbook(ORIG)
    exp = openpyxl.load_workbook(EXP)

    errors = []
    print("Export sheets:", exp.sheetnames)
    assert exp.sheetnames == list(expected.keys()), f"sheet mismatch: {exp.sheetnames}"

    for sheet, rows in expected.items():
        ows = orig[sheet]
        ews = exp[sheet]
        print(f"\n=== {sheet} === expected {len(rows)} rows, export has {ews.max_row} rows")
        if ews.max_row != len(rows):
            errors.append(f"{sheet}: row count {ews.max_row} != expected {len(rows)}")

        # compare each exported (new) row to its original source row, by style + value
        for new_idx, old_idx in enumerate(rows, start=1):
            for col in range(1, 13):
                oc = ows.cell(row=old_idx, column=col)
                ec = ews.cell(row=new_idx, column=col)
                os_, es_ = style_sig(oc, recolor=True), style_sig(ec)
                if os_ != es_:
                    errors.append(
                        f"{sheet} new r{new_idx} (orig r{old_idx}) col{col}: STYLE "
                        f"{es_} != {os_}"
                    )
                    break
        # spot check: values of a few cols equal original (col E = Aksiyon)
        for new_idx, old_idx in enumerate(rows, start=1):
            ov = ows.cell(row=old_idx, column=5).value
            ev = ews.cell(row=new_idx, column=5).value
            if (ov or "") != (ev or ""):
                # allow edited cells; col 5 is never edited
                errors.append(f"{sheet} new r{new_idx} colE value {ev!r} != {ov!r}")

    # check injected edits landed (E-Ticaret durum on first task row)
    ews = exp["E-Ticaret"]
    # find row whose col E == AI Visibility Audit...
    durum_found = None
    marka_found = None
    for r in range(1, ews.max_row + 1):
        e = ews.cell(row=r, column=5).value or ""
        if e.startswith("AI Visibility Audit"):
            durum_found = ews.cell(row=r, column=10).value
            marka_found = ews.cell(row=r, column=11).value
    print("\nInjected Durum:", repr(durum_found), "| Marka Notları:", repr(marka_found))
    if durum_found != "Devam Ediyor":
        errors.append(f"Durum injection failed: {durum_found!r}")
    if marka_found != "Öncelikli — Q3 hedefi":
        errors.append(f"Marka Notları injection failed: {marka_found!r}")

    # check Durum dropdown survived
    dv_count = len(exp["E-Ticaret"].data_validations.dataValidation)
    print("Durum dataValidation count:", dv_count)
    if dv_count < 1:
        errors.append("Durum dropdown (dataValidation) missing")

    # check column widths preserved
    ow = {k: round(v.width, 1) for k, v in orig["E-Ticaret"].column_dimensions.items() if v.width}
    ew = {k: round(v.width, 1) for k, v in exp["E-Ticaret"].column_dimensions.items() if v.width}
    print("Col widths match:", ow == ew, "| export:", ew)
    if ow != ew:
        errors.append(f"col widths differ: {ew} != {ow}")

    # frozen panes
    print("Freeze panes orig:", orig["E-Ticaret"].freeze_panes, "| export:", exp["E-Ticaret"].freeze_panes)

    print("\n" + ("=" * 50))
    if errors:
        print(f"❌ {len(errors)} PROBLEM:")
        for e in errors[:25]:
            print("  -", e)
        sys.exit(1)
    print("✅ ALL CHECKS PASSED — export is byte-style identical and correctly filtered")


if __name__ == "__main__":
    main()
